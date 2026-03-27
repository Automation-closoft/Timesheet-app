from pathlib import Path
import os
import logging
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.utils import timezone
from .models import UserProfile
from .forms import UserRegistrationForm
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
import openpyxl
from django.http import FileResponse, Http404
from urllib.parse import unquote

# Set up logging
logger = logging.getLogger(__name__)

# Path to save Excel files
EXCEL_PATH = 'timesheets/'

# Ensure the path exists
if not os.path.exists(EXCEL_PATH):
    os.makedirs(EXCEL_PATH)


def format_hours_and_minutes(total_hours):
    hours = int(total_hours)
    minutes = int((total_hours - hours) * 60)
    return f"{hours}h {minutes}m"


@login_required
def signup(request):
    if not request.user.is_staff:
        raise PermissionDenied("You do not have permission to create new users.")

    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            employee_name = form.cleaned_data.get('employee_name')

            user = User.objects.create_user(username=username, password=password)
            profile = UserProfile.objects.create(user=user, employee_name=employee_name)

            # Create a new Excel sheet for the user
            excel_filename = os.path.join(EXCEL_PATH, f'{employee_name}.xlsx')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['Date', 'Project Working On', 'Log In Time', 'Log Out Time', 'Hours Worked'])
            wb.save(excel_filename)

            return redirect('login')
    else:
        form = UserRegistrationForm()
    return render(request, 'signup.html', {'form': form})


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            return render(request, 'login.html', {'error': 'Invalid credentials'})
    return render(request, 'login.html')


@login_required
def home(request):
    current_date = timezone.now().date()

    if request.method == 'POST':
        project = request.POST.get('project')
        date = request.POST.get('date')
        login_time = request.POST.get('login_time')
        logout_time = request.POST.get('logout_time')

        if not all([project, date, login_time, logout_time]):
            messages.error(request, "Please fill in all fields.")
            return render(request, 'home.html', {'current_date': current_date})

        try:
            login_time_obj = datetime.strptime(login_time, '%H:%M')
            logout_time_obj = datetime.strptime(logout_time, '%H:%M')
        except ValueError:
            messages.error(request, "Please enter time in HH:MM format.")
            return render(request, 'home.html', {'current_date': current_date})

        hours_worked = (logout_time_obj - login_time_obj).seconds / 3600
        formatted_hours_worked = format_hours_and_minutes(hours_worked)

        try:
            profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
            messages.error(request, "User profile does not exist.")
            return redirect('signup')

        excel_filename = os.path.join(EXCEL_PATH, f'{profile.employee_name}.xlsx')

        if not os.path.exists(excel_filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['Date', 'Project Working On', 'Log In Time', 'Log Out Time', 'Hours Worked'])
        else:
            wb = openpyxl.load_workbook(excel_filename)
            ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value == date:
                row[1].value = project
                row[2].value = login_time
                row[3].value = logout_time
                row[4].value = formatted_hours_worked
                break
        else:
            ws.append([date, project, login_time, logout_time, formatted_hours_worked])

        wb.save(excel_filename)

        return redirect('success')

    return render(request, 'home.html', {'current_date': current_date})


@login_required
def admin_download_timesheets(request):
    if not request.user.is_staff:
        raise PermissionDenied("You do not have permission to download timesheets.")

    # If ?file=something.xlsx is in the URL, download that file
    file_to_download = request.GET.get('file')
    if file_to_download:
        # Decode URL-encoded characters (e.g., %20 -> space)
        file_to_download = unquote(file_to_download)

        # Prevent path traversal attacks
        if '..' in file_to_download or '/' in file_to_download:
            raise Http404("Invalid filename")

        excel_filepath = os.path.join(EXCEL_PATH, file_to_download)

        if os.path.exists(excel_filepath):
            response = FileResponse(
                open(excel_filepath, 'rb'),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{file_to_download}"'
            return response
        else:
            raise Http404("Timesheet not found")

    # Otherwise, show the list of timesheets
    profiles = UserProfile.objects.all()
    excel_files = []
    for profile in profiles:
        excel_filename = os.path.join(EXCEL_PATH, f'{profile.employee_name}.xlsx')
        if os.path.exists(excel_filename):
            filename = f'{profile.employee_name}.xlsx'
            excel_files.append({
                'name': f"{profile.employee_name}'s Timesheet",
                'url': f'/download-timesheets/?file={filename}'
            })

    return render(request, 'download_timesheets.html', {'excel_files': excel_files})


def logout_view(request):
    logout(request)
    return redirect('login')


def success_view(request):
    return render(request, 'success.html')


@login_required
def password_change_view(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()

            storage = messages.get_messages(request)
            for _ in storage:
                pass

            messages.success(request, 'Your password has been updated!')

            logout(request)
            return redirect('password_change_done')
        else:
            logger.error(f"Password change form errors: {form.errors}")
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'password_change_form.html', {'form': form})


def password_change_done(request):
    return render(request, 'password_change_done.html')
